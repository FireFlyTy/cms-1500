#!/usr/bin/env python3
"""
Load reference data from xlsx/csv files into SQLite database
"""
import sys
import re
import csv
from pathlib import Path

# Add project root to path (api/scripts/ -> api/ -> project root)
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from config import (
    HCPCS_DIR, NCCI_DIR, ICD10_DIR, CPT_DIR,
    REFERENCE_DB_PATH, EXPECTED_FILES
)
import pandas as pd
from src.db.models import init_db

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_hcpcs_codes(conn, filepath: Path):
    """Load HCPCS codes from xlsx"""
    print(f"Loading HCPCS codes from {filepath.name}...")
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb.active
    
    # Get header row
    headers = [cell.value for cell in sheet[1]]
    
    # Map columns
    col_map = {
        "code": headers.index("HCPC") if "HCPC" in headers else 0,
        "long_desc": headers.index("LONG DESCRIPTION") if "LONG DESCRIPTION" in headers else 3,
        "short_desc": headers.index("SHORT DESCRIPTION") if "SHORT DESCRIPTION" in headers else 4,
        "betos": headers.index("BETOS") if "BETOS" in headers else 37,
        "tos": headers.index("TOS1") if "TOS1" in headers else 38,
        "coverage": headers.index("COV") if "COV" in headers else 30,
        "proc_note": headers.index("PROCNOTE") if "PROCNOTE" in headers else 36,
        "add_date": headers.index("ADD DT") if "ADD DT" in headers else 44,
        "term_date": headers.index("TERM DT") if "TERM DT" in headers else 46,
    }
    
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else None
        if not code:
            continue
            
        conn.execute("""
            INSERT OR REPLACE INTO hcpcs 
            (code, long_description, short_description, betos, tos, coverage, proc_note, add_date, term_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            code,
            row[col_map["long_desc"]],
            row[col_map["short_desc"]],
            row[col_map["betos"]],
            row[col_map["tos"]],
            row[col_map["coverage"]],
            row[col_map["proc_note"]],
            row[col_map["add_date"]],
            row[col_map["term_date"]],
        ))
        count += 1
        if count % 1000 == 0:
            print(f"  Loaded {count} codes...")
    
    conn.commit()
    print(f"  ✓ Loaded {count} HCPCS codes")
    return count


def load_hcpcs_notes(conn, filepath: Path):
    """Load HCPCS processing notes from txt file"""
    print(f"Loading HCPCS notes from {filepath.name}...")
    
    with open(filepath, "r", encoding="latin-1") as f:
        content = f.read()
    
    # Parse notes: pattern is "NNNN--text"
    notes = {}
    current_note = None
    current_text = []
    
    for line in content.split('\n'):
        match = re.match(r'\s+(\d{4})--(.+)', line)
        if match:
            if current_note:
                notes[current_note] = ' '.join(current_text).strip().rstrip('*').strip()
            current_note = match.group(1)
            current_text = [match.group(2).strip().rstrip('*').strip()]
        elif current_note and line.strip() and not line.strip().startswith('*'):
            text = line.strip().rstrip('*').strip()
            if text:
                current_text.append(text)
    
    if current_note:
        notes[current_note] = ' '.join(current_text).strip()
    
    count = 0
    for note_id, note_text in notes.items():
        conn.execute("""
            INSERT OR REPLACE INTO hcpcs_notes (note_id, note_text)
            VALUES (?, ?)
        """, (note_id, note_text))
        count += 1
    
    conn.commit()
    print(f"  ✓ Loaded {count} HCPCS notes")
    return count


def load_ncci_ptp(conn, filepath: Path):
    """Load NCCI PTP edits from xlsx"""
    print(f"Loading NCCI PTP from {filepath.name}...")
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb.active
    
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        
        col1 = str(row[0]).strip()
        col2 = str(row[1]).strip()
        
        # Skip header rows
        if col1 == "Column 1" or "copyright" in col1.lower():
            continue
        
        # Parse modifier indicator
        mod_ind = row[5] if len(row) > 5 else None
        if mod_ind is not None:
            try:
                mod_ind = int(mod_ind)
            except (ValueError, TypeError):
                mod_ind = None
        
        conn.execute("""
            INSERT OR IGNORE INTO ncci_ptp 
            (column1, column2, modifier_indicator, effective_date, deletion_date, rationale)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            col1,
            col2,
            mod_ind,
            row[3] if len(row) > 3 else None,  # effective date
            row[4] if len(row) > 4 else None,  # deletion date
            row[6] if len(row) > 6 else None,  # rationale
        ))
        count += 1
        if count % 50000 == 0:
            print(f"  Loaded {count} PTP edits...")
    
    conn.commit()
    print(f"  ✓ Loaded {count} NCCI PTP edits")
    return count


def load_ncci_mue(conn, filepath: Path, table_name: str):
    """Load NCCI MUE edits from csv"""
    print(f"Loading NCCI MUE ({table_name}) from {filepath.name}...")
    
    count = 0
    with open(filepath, "r", encoding="latin-1") as f:
        reader = csv.reader(f)
        
        for row in reader:
            if not row or len(row) < 2:
                continue
            
            code = row[0].strip()
            
            # Skip header and copyright rows
            if not code or code.startswith('"') or "HCPCS" in code or "CPT" in code:
                continue
            
            # Parse MUE value
            try:
                mue_value = int(row[1])
            except (ValueError, IndexError):
                continue
            
            conn.execute(f"""
                INSERT OR REPLACE INTO {table_name}
                (code, mue_value, adjudication_indicator, rationale)
                VALUES (?, ?, ?, ?)
            """, (
                code,
                mue_value,
                row[2] if len(row) > 2 else None,
                row[3] if len(row) > 3 else None,
            ))
            count += 1
    
    conn.commit()
    print(f"  ✓ Loaded {count} MUE edits into {table_name}")
    return count


def load_icd10_codes(conn, filepath: Path):
    """Load ICD-10 codes from csv"""
    print(f"Loading ICD-10 codes from {filepath.name}...")

    count = 0
    with open(filepath, "r", encoding="utf-8-sig") as f:
        # Detect delimiter
        sample = f.read(1024)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample)
            reader = csv.DictReader(f, delimiter=dialect.delimiter)
        except csv.Error:
            reader = csv.DictReader(f)

        for row in reader:
            code = row.get("code") or row.get("Code") or row.get("CODE")
            desc = row.get("description") or row.get("Description") or row.get("DESCRIPTION")

            if code:
                conn.execute("""
                    INSERT OR REPLACE INTO icd10 (code, description)
                    VALUES (?, ?)
                """, (code.replace(".", "").upper(), desc))
                count += 1

    conn.commit()
    print(f"  ✓ Loaded {count} ICD-10 codes")
    return count


def load_cpt_rvu(conn, filepath: Path):
    """Load CPT codes from CMS RVU file"""
    print(f"Loading CPT codes from {filepath.name}...")

    # Read CSV, skip header rows
    df = pd.read_csv(filepath, skiprows=9, low_memory=False)

    # Get code and description
    df_codes = df[['HCPCS', 'DESCRIPTION', 'CODE']].dropna(subset=['HCPCS'])
    df_codes = df_codes[df_codes['HCPCS'].astype(str).str.strip() != '']
    df_codes.columns = ['code', 'description', 'status']
    df_codes['code'] = df_codes['code'].astype(str).str.strip()
    df_codes['description'] = df_codes['description'].astype(str).str.strip()
    df_codes = df_codes.drop_duplicates(subset=['code'], keep='first')

    count = 0
    for _, row in df_codes.iterrows():
        conn.execute("""
            INSERT OR REPLACE INTO cpt (code, description, status)
            VALUES (?, ?, ?)
        """, (row['code'], row['description'], row['status']))
        count += 1
        if count % 5000 == 0:
            print(f"  Loaded {count} codes...")

    conn.commit()
    print(f"  ✓ Loaded {count} CPT codes from RVU")
    return count


def load_cpt_dhs_addendum(conn, filepath: Path):
    """Load additional CPT codes from DHS addendum"""
    print(f"Loading CPT codes from {filepath.name}...")

    df = pd.read_excel(filepath, header=None)

    # Extract codes (code in column 0, description in column 1)
    count = 0
    for idx, row in df.iterrows():
        val = str(row[0]).strip()
        # Check if it looks like a CPT code (4-6 chars, starts with digit or letter)
        if len(val) >= 4 and len(val) <= 6 and (val[0].isdigit() or val[0] in 'AEJGHQST'):
            desc = str(row[1]).strip() if pd.notna(row[1]) else ''
            if desc and desc != 'nan':
                conn.execute("""
                    INSERT INTO cpt (code, description)
                    VALUES (?, ?)
                    ON CONFLICT(code) DO UPDATE SET
                        description = COALESCE(NULLIF(cpt.description, ''), excluded.description)
                """, (val, desc))
                count += 1

    conn.commit()
    print(f"  ✓ Loaded {count} additional CPT codes from DHS addendum")
    return count


def build_code_hierarchy(conn):
    """
    Build code_hierarchy table from existing ICD-10, CPT, HCPCS codes.

    Hierarchy levels:
    - Level 0 (Meta-category): E, F, 9, J (first letter/digit)
    - Level 1 (Category): E11, F32, 992 (3-char ICD-10, 3-digit CPT prefix)
    - Level 2 (Subcategory): E11.6, 9921 (intermediate grouping)
    - Level 3 (Code): E11.65, 99213 (full specific codes)

    Meta-categories:
    - ICD-10: First letter (A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, V, W, X, Y, Z)
    - HCPCS: First letter (A, B, C, E, G, J, K, L, M, P, Q, R, S, T, V)
    - CPT: First digit (0, 1, 2, 3, 4, 5, 6, 7, 8, 9)
    """
    print("\nBuilding code hierarchy...")

    cursor = conn.cursor()

    # Clear existing hierarchy
    cursor.execute("DELETE FROM code_hierarchy")

    meta_categories_added = set()
    categories_added = set()
    subcategories_added = set()
    codes_added = 0

    # ============================================================
    # ICD-10 Hierarchy
    # ============================================================
    print("  Processing ICD-10 codes...")

    cursor.execute("SELECT code, description FROM icd10")
    icd10_rows = cursor.fetchall()

    for row in icd10_rows:
        code_raw = row[0]  # Stored without dots (E119, F321)
        description = row[1]

        # Skip if too short
        if len(code_raw) < 3:
            continue

        # Format code with dot for ICD-10 (E119 -> E11.9)
        if len(code_raw) > 3:
            code = code_raw[:3] + '.' + code_raw[3:]
        else:
            code = code_raw

        # Meta-category: first letter (E, F, G, etc.)
        meta_cat = code_raw[0].upper()
        meta_key = ('ICD-10', meta_cat)

        # Level 0: Meta-category (single letter, e.g., E)
        if meta_key not in meta_categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('ICD-10', 0, ?, NULL, ?, ?)
            """, (meta_cat, f"ICD-10 Chapter {meta_cat}", meta_cat))
            meta_categories_added.add(meta_key)

        # Level 1: Category (first 3 chars, e.g., E11)
        category = code_raw[:3]
        category_key = ('ICD-10', category)

        if category_key not in categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('ICD-10', 1, ?, ?, ?, ?)
            """, (category, meta_cat, f"ICD-10 Category {category}", meta_cat))
            categories_added.add(category_key)

        # Level 2: Subcategory (e.g., E11.6 from E11.65)
        if '.' in code and len(code.split('.')[1]) > 1:
            parts = code.split('.')
            subcategory = parts[0] + '.' + parts[1][0]
            subcat_key = ('ICD-10', subcategory)

            if subcat_key not in subcategories_added:
                cursor.execute("""
                    INSERT OR IGNORE INTO code_hierarchy
                    (code_type, level, pattern, parent_pattern, description, meta_category)
                    VALUES ('ICD-10', 2, ?, ?, ?, ?)
                """, (subcategory, category, f"ICD-10 Subcategory {subcategory}", meta_cat))
                subcategories_added.add(subcat_key)

            parent = subcategory
        else:
            parent = category

        # Level 3: Full code
        cursor.execute("""
            INSERT OR IGNORE INTO code_hierarchy
            (code_type, level, pattern, parent_pattern, description, meta_category)
            VALUES ('ICD-10', 3, ?, ?, ?, ?)
        """, (code, parent, description, meta_cat))
        codes_added += 1

    print(f"    ICD-10: {len([c for c in meta_categories_added if c[0]=='ICD-10'])} meta-categories, "
          f"{len([c for c in categories_added if c[0]=='ICD-10'])} categories, "
          f"{len([c for c in subcategories_added if c[0]=='ICD-10'])} subcategories")

    # ============================================================
    # CPT Hierarchy
    # ============================================================
    print("  Processing CPT codes...")

    cursor.execute("SELECT code, description FROM cpt")
    cpt_rows = cursor.fetchall()

    for row in cpt_rows:
        code = str(row[0]).strip()
        description = row[1]

        # CPT codes are typically 5 digits
        if len(code) < 4:
            continue

        # Meta-category: first digit (0-9) representing CPT section
        meta_cat = code[0] if code[0].isdigit() else '9'
        meta_key = ('CPT', meta_cat)

        # Level 0: Meta-category (single digit, e.g., 9)
        if meta_key not in meta_categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('CPT', 0, ?, NULL, ?, ?)
            """, (meta_cat, f"CPT Section {meta_cat}xxxx", meta_cat))
            meta_categories_added.add(meta_key)

        # Level 1: Category (first 3 digits, e.g., 992)
        category = code[:3]
        category_key = ('CPT', category)

        if category_key not in categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('CPT', 1, ?, ?, ?, ?)
            """, (category, meta_cat, f"CPT Category {category}xx", meta_cat))
            categories_added.add(category_key)

        # Level 2: Subcategory (first 4 digits, e.g., 9921)
        if len(code) >= 5:
            subcategory = code[:4]
            subcat_key = ('CPT', subcategory)

            if subcat_key not in subcategories_added:
                cursor.execute("""
                    INSERT OR IGNORE INTO code_hierarchy
                    (code_type, level, pattern, parent_pattern, description, meta_category)
                    VALUES ('CPT', 2, ?, ?, ?, ?)
                """, (subcategory, category, f"CPT Subcategory {subcategory}x", meta_cat))
                subcategories_added.add(subcat_key)

            parent = subcategory
        else:
            parent = category

        # Level 3: Full code
        cursor.execute("""
            INSERT OR IGNORE INTO code_hierarchy
            (code_type, level, pattern, parent_pattern, description, meta_category)
            VALUES ('CPT', 3, ?, ?, ?, ?)
        """, (code, parent, description, meta_cat))
        codes_added += 1

    print(f"    CPT: {len([c for c in meta_categories_added if c[0]=='CPT'])} meta-categories, "
          f"{len([c for c in categories_added if c[0]=='CPT'])} categories, "
          f"{len([c for c in subcategories_added if c[0]=='CPT'])} subcategories")

    # ============================================================
    # HCPCS Hierarchy
    # ============================================================
    print("  Processing HCPCS codes...")

    cursor.execute("SELECT code, long_description FROM hcpcs")
    hcpcs_rows = cursor.fetchall()

    for row in hcpcs_rows:
        code = str(row[0]).strip()
        description = row[1]

        # HCPCS codes start with letter (A, E, G, J, etc.)
        if len(code) < 4 or not code[0].isalpha():
            continue

        # Meta-category: first letter (A, B, C, E, G, J, etc.)
        meta_cat = code[0].upper()
        meta_key = ('HCPCS', meta_cat)

        # Level 0: Meta-category (single letter, e.g., J)
        if meta_key not in meta_categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('HCPCS', 0, ?, NULL, ?, ?)
            """, (meta_cat, f"HCPCS Section {meta_cat}xxxx", meta_cat))
            meta_categories_added.add(meta_key)

        # Level 1: Category (first 2 chars, e.g., J1, A4, E0)
        category = code[:2]
        category_key = ('HCPCS', category)

        if category_key not in categories_added:
            cursor.execute("""
                INSERT OR IGNORE INTO code_hierarchy
                (code_type, level, pattern, parent_pattern, description, meta_category)
                VALUES ('HCPCS', 1, ?, ?, ?, ?)
            """, (category, meta_cat, f"HCPCS Category {category}xxx", meta_cat))
            categories_added.add(category_key)

        # Level 2: Subcategory (first 3 chars, e.g., J19, A42)
        if len(code) >= 4:
            subcategory = code[:3]
            subcat_key = ('HCPCS', subcategory)

            if subcat_key not in subcategories_added:
                cursor.execute("""
                    INSERT OR IGNORE INTO code_hierarchy
                    (code_type, level, pattern, parent_pattern, description, meta_category)
                    VALUES ('HCPCS', 2, ?, ?, ?, ?)
                """, (subcategory, category, f"HCPCS Subcategory {subcategory}xx", meta_cat))
                subcategories_added.add(subcat_key)

            parent = subcategory
        else:
            parent = category

        # Level 3: Full code
        cursor.execute("""
            INSERT OR IGNORE INTO code_hierarchy
            (code_type, level, pattern, parent_pattern, description, meta_category)
            VALUES ('HCPCS', 3, ?, ?, ?, ?)
        """, (code, parent, description, meta_cat))
        codes_added += 1

    print(f"    HCPCS: {len([c for c in meta_categories_added if c[0]=='HCPCS'])} meta-categories, "
          f"{len([c for c in categories_added if c[0]=='HCPCS'])} categories, "
          f"{len([c for c in subcategories_added if c[0]=='HCPCS'])} subcategories")

    conn.commit()

    # Get final counts
    cursor.execute("SELECT level, COUNT(*) FROM code_hierarchy GROUP BY level")
    level_counts = {row[0]: row[1] for row in cursor.fetchall()}

    print(f"\n  ✓ Code hierarchy built:")
    print(f"    Level 0 (Meta-categories): {level_counts.get(0, 0):,}")
    print(f"    Level 1 (Categories): {level_counts.get(1, 0):,}")
    print(f"    Level 2 (Subcategories): {level_counts.get(2, 0):,}")
    print(f"    Level 3 (Codes): {level_counts.get(3, 0):,}")

    return sum(level_counts.values())


def main():
    print("=" * 60)
    print("Loading Reference Data into SQLite")
    print("=" * 60)
    
    # Initialize database
    print(f"\nDatabase: {REFERENCE_DB_PATH}")
    conn = init_db(REFERENCE_DB_PATH)
    
    stats = {}
    
    # Load HCPCS
    hcpcs_codes_path = HCPCS_DIR / EXPECTED_FILES["hcpcs"]["codes"]
    hcpcs_notes_path = HCPCS_DIR / EXPECTED_FILES["hcpcs"]["notes"]
    
    if hcpcs_codes_path.exists():
        stats["hcpcs_codes"] = load_hcpcs_codes(conn, hcpcs_codes_path)
    else:
        print(f"⚠ HCPCS codes not found: {hcpcs_codes_path}")
    
    if hcpcs_notes_path.exists():
        stats["hcpcs_notes"] = load_hcpcs_notes(conn, hcpcs_notes_path)
    else:
        print(f"⚠ HCPCS notes not found: {hcpcs_notes_path}")
    
    # Load NCCI PTP
    ncci_ptp_path = NCCI_DIR / EXPECTED_FILES["ncci"]["ptp_pra"]
    if ncci_ptp_path.exists():
        stats["ncci_ptp"] = load_ncci_ptp(conn, ncci_ptp_path)
    else:
        print(f"⚠ NCCI PTP not found: {ncci_ptp_path}")
    
    # Load NCCI MUE
    mue_pra_path = NCCI_DIR / EXPECTED_FILES["ncci"]["mue_pra"]
    if mue_pra_path.exists():
        stats["ncci_mue_pra"] = load_ncci_mue(conn, mue_pra_path, "ncci_mue_pra")
    else:
        print(f"⚠ NCCI MUE (PRA) not found: {mue_pra_path}")
    
    mue_dme_path = NCCI_DIR / EXPECTED_FILES["ncci"]["mue_dme"]
    if mue_dme_path.exists():
        stats["ncci_mue_dme"] = load_ncci_mue(conn, mue_dme_path, "ncci_mue_dme")
    else:
        print(f"⚠ NCCI MUE (DME) not found: {mue_dme_path}")
    
    # Load ICD-10
    icd10_path = ICD10_DIR / EXPECTED_FILES["icd10"]["codes"]
    if icd10_path.exists():
        stats["icd10"] = load_icd10_codes(conn, icd10_path)
    else:
        print(f"⚠ ICD-10 codes not found: {icd10_path}")

    # Load CPT from RVU file
    cpt_rvu_path = CPT_DIR / EXPECTED_FILES["cpt"]["rvu"]
    if cpt_rvu_path.exists():
        stats["cpt_rvu"] = load_cpt_rvu(conn, cpt_rvu_path)
    else:
        print(f"⚠ CPT RVU file not found: {cpt_rvu_path}")

    # Load CPT from DHS addendum
    cpt_dhs_path = CPT_DIR / EXPECTED_FILES["cpt"]["dhs_addendum"]
    if cpt_dhs_path.exists():
        stats["cpt_dhs"] = load_cpt_dhs_addendum(conn, cpt_dhs_path)
    else:
        print(f"⚠ CPT DHS addendum not found: {cpt_dhs_path}")

    # Build code hierarchy from loaded codes
    stats["code_hierarchy"] = build_code_hierarchy(conn)

    conn.close()
    
    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for key, count in stats.items():
        print(f"  {key}: {count:,} records")
    print(f"\nDatabase saved: {REFERENCE_DB_PATH}")


if __name__ == "__main__":
    main()
