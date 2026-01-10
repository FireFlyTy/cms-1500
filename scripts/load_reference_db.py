#!/usr/bin/env python3
"""
Load reference data from xlsx/csv files into SQLite database
"""
import sys
import re
import csv
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import (
    HCPCS_DIR, NCCI_DIR, ICD10_DIR,
    REFERENCE_DB_PATH, EXPECTED_FILES
)
from src.db.models import init_db

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_hcpcs_codes(conn, filepath: Path):
    """Load HCPCS codes from xlsx"""
    print(f"Loading HCPCS codes from {filepath.name}...")
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb.active
    
    # Get header row
    headers = [cell.value for cell in sheet[1]]
    
    # Map columns
    col_map = {
        "code": headers.index("HCPC") if "HCPC" in headers else 0,
        "long_desc": headers.index("LONG DESCRIPTION") if "LONG DESCRIPTION" in headers else 3,
        "short_desc": headers.index("SHORT DESCRIPTION") if "SHORT DESCRIPTION" in headers else 4,
        "betos": headers.index("BETOS") if "BETOS" in headers else 37,
        "tos": headers.index("TOS1") if "TOS1" in headers else 38,
        "coverage": headers.index("COV") if "COV" in headers else 30,
        "proc_note": headers.index("PROCNOTE") if "PROCNOTE" in headers else 36,
        "add_date": headers.index("ADD DT") if "ADD DT" in headers else 44,
        "term_date": headers.index("TERM DT") if "TERM DT" in headers else 46,
    }
    
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else None
        if not code:
            continue
            
        conn.execute("""
            INSERT OR REPLACE INTO hcpcs 
            (code, long_description, short_description, betos, tos, coverage, proc_note, add_date, term_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            code,
            row[col_map["long_desc"]],
            row[col_map["short_desc"]],
            row[col_map["betos"]],
            row[col_map["tos"]],
            row[col_map["coverage"]],
            row[col_map["proc_note"]],
            row[col_map["add_date"]],
            row[col_map["term_date"]],
        ))
        count += 1
        if count % 1000 == 0:
            print(f"  Loaded {count} codes...")
    
    conn.commit()
    print(f"  ✓ Loaded {count} HCPCS codes")
    return count


def load_hcpcs_notes(conn, filepath: Path):
    """Load HCPCS processing notes from txt file"""
    print(f"Loading HCPCS notes from {filepath.name}...")
    
    with open(filepath, "r", encoding="latin-1") as f:
        content = f.read()
    
    # Parse notes: pattern is "NNNN--text"
    notes = {}
    current_note = None
    current_text = []
    
    for line in content.split('\n'):
        match = re.match(r'\s+(\d{4})--(.+)', line)
        if match:
            if current_note:
                notes[current_note] = ' '.join(current_text).strip().rstrip('*').strip()
            current_note = match.group(1)
            current_text = [match.group(2).strip().rstrip('*').strip()]
        elif current_note and line.strip() and not line.strip().startswith('*'):
            text = line.strip().rstrip('*').strip()
            if text:
                current_text.append(text)
    
    if current_note:
        notes[current_note] = ' '.join(current_text).strip()
    
    count = 0
    for note_id, note_text in notes.items():
        conn.execute("""
            INSERT OR REPLACE INTO hcpcs_notes (note_id, note_text)
            VALUES (?, ?)
        """, (note_id, note_text))
        count += 1
    
    conn.commit()
    print(f"  ✓ Loaded {count} HCPCS notes")
    return count


def load_ncci_ptp(conn, filepath: Path):
    """Load NCCI PTP edits from xlsx"""
    print(f"Loading NCCI PTP from {filepath.name}...")
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb.active
    
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        
        col1 = str(row[0]).strip()
        col2 = str(row[1]).strip()
        
        # Skip header rows
        if col1 == "Column 1" or "copyright" in col1.lower():
            continue
        
        # Parse modifier indicator
        mod_ind = row[5] if len(row) > 5 else None
        if mod_ind is not None:
            try:
                mod_ind = int(mod_ind)
            except (ValueError, TypeError):
                mod_ind = None
        
        conn.execute("""
            INSERT OR IGNORE INTO ncci_ptp 
            (column1, column2, modifier_indicator, effective_date, deletion_date, rationale)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            col1,
            col2,
            mod_ind,
            row[3] if len(row) > 3 else None,  # effective date
            row[4] if len(row) > 4 else None,  # deletion date
            row[6] if len(row) > 6 else None,  # rationale
        ))
        count += 1
        if count % 50000 == 0:
            print(f"  Loaded {count} PTP edits...")
    
    conn.commit()
    print(f"  ✓ Loaded {count} NCCI PTP edits")
    return count


def load_ncci_mue(conn, filepath: Path, table_name: str):
    """Load NCCI MUE edits from csv"""
    print(f"Loading NCCI MUE ({table_name}) from {filepath.name}...")
    
    count = 0
    with open(filepath, "r", encoding="latin-1") as f:
        reader = csv.reader(f)
        
        for row in reader:
            if not row or len(row) < 2:
                continue
            
            code = row[0].strip()
            
            # Skip header and copyright rows
            if not code or code.startswith('"') or "HCPCS" in code or "CPT" in code:
                continue
            
            # Parse MUE value
            try:
                mue_value = int(row[1])
            except (ValueError, IndexError):
                continue
            
            conn.execute(f"""
                INSERT OR REPLACE INTO {table_name}
                (code, mue_value, adjudication_indicator, rationale)
                VALUES (?, ?, ?, ?)
            """, (
                code,
                mue_value,
                row[2] if len(row) > 2 else None,
                row[3] if len(row) > 3 else None,
            ))
            count += 1
    
    conn.commit()
    print(f"  ✓ Loaded {count} MUE edits into {table_name}")
    return count


def load_icd10_codes(conn, filepath: Path):
    """Load ICD-10 codes from csv"""
    print(f"Loading ICD-10 codes from {filepath.name}...")
    
    count = 0
    with open(filepath, "r", encoding="utf-8-sig") as f:
        # Detect delimiter
        sample = f.read(1024)
        f.seek(0)
        
        try:
            dialect = csv.Sniffer().sniff(sample)
            reader = csv.DictReader(f, delimiter=dialect.delimiter)
        except csv.Error:
            reader = csv.DictReader(f)
        
        for row in reader:
            code = row.get("code") or row.get("Code") or row.get("CODE")
            desc = row.get("description") or row.get("Description") or row.get("DESCRIPTION")
            
            if code:
                conn.execute("""
                    INSERT OR REPLACE INTO icd10 (code, description)
                    VALUES (?, ?)
                """, (code.replace(".", "").upper(), desc))
                count += 1
    
    conn.commit()
    print(f"  ✓ Loaded {count} ICD-10 codes")
    return count


def main():
    print("=" * 60)
    print("Loading Reference Data into SQLite")
    print("=" * 60)
    
    # Initialize database
    print(f"\nDatabase: {REFERENCE_DB_PATH}")
    conn = init_db(REFERENCE_DB_PATH)
    
    stats = {}
    
    # Load HCPCS
    hcpcs_codes_path = HCPCS_DIR / EXPECTED_FILES["hcpcs"]["codes"]
    hcpcs_notes_path = HCPCS_DIR / EXPECTED_FILES["hcpcs"]["notes"]
    
    if hcpcs_codes_path.exists():
        stats["hcpcs_codes"] = load_hcpcs_codes(conn, hcpcs_codes_path)
    else:
        print(f"⚠ HCPCS codes not found: {hcpcs_codes_path}")
    
    if hcpcs_notes_path.exists():
        stats["hcpcs_notes"] = load_hcpcs_notes(conn, hcpcs_notes_path)
    else:
        print(f"⚠ HCPCS notes not found: {hcpcs_notes_path}")
    
    # Load NCCI PTP
    ncci_ptp_path = NCCI_DIR / EXPECTED_FILES["ncci"]["ptp_pra"]
    if ncci_ptp_path.exists():
        stats["ncci_ptp"] = load_ncci_ptp(conn, ncci_ptp_path)
    else:
        print(f"⚠ NCCI PTP not found: {ncci_ptp_path}")
    
    # Load NCCI MUE
    mue_pra_path = NCCI_DIR / EXPECTED_FILES["ncci"]["mue_pra"]
    if mue_pra_path.exists():
        stats["ncci_mue_pra"] = load_ncci_mue(conn, mue_pra_path, "ncci_mue_pra")
    else:
        print(f"⚠ NCCI MUE (PRA) not found: {mue_pra_path}")
    
    mue_dme_path = NCCI_DIR / EXPECTED_FILES["ncci"]["mue_dme"]
    if mue_dme_path.exists():
        stats["ncci_mue_dme"] = load_ncci_mue(conn, mue_dme_path, "ncci_mue_dme")
    else:
        print(f"⚠ NCCI MUE (DME) not found: {mue_dme_path}")
    
    # Load ICD-10
    icd10_path = ICD10_DIR / EXPECTED_FILES["icd10"]["codes"]
    if icd10_path.exists():
        stats["icd10"] = load_icd10_codes(conn, icd10_path)
    else:
        print(f"⚠ ICD-10 codes not found: {icd10_path}")
    
    conn.close()
    
    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for key, count in stats.items():
        print(f"  {key}: {count:,} records")
    print(f"\nDatabase saved: {REFERENCE_DB_PATH}")


if __name__ == "__main__":
    main()
